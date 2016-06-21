#! /usr/bin/env python3
# -*- coding:utf-8 -*-
'''
Created on 2016年6月13日

@author: GiliGilieye
'''
from openpyxl import Workbook
from openpyxl import load_workbook
import warnings


class ExcelSGH():
    def __init__(self):
        warnings.simplefilter("ignore")
        self.list_a = self.load_a()
        self.load_b()
        self.c = self.load_c()
        self.main()

    def load_a(self):
        list_a = []
        dict_a = {}
        list_room = []
        list_name = []
        # 这里是load一个xlsx文件
        wb = load_workbook(filename=r'../a.xlsx')
        # ws = wb.get_sheet_names()
        # 上面一句是取xlsx文件里面所有的Sheet名,return一个list
        # 上一句的输出是['表名1', '表名2']
        # 这里是取到文件a第一个Sheet里面的内容
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        # 以下两句是取到汇总这个Sheet里的行数和字段数（就是列数）
        # 一开始我用教程上的这两句，会提示Use Warning，是因为这两个方法不再被推荐使用了，他推荐使用以下方法来取到行列数。
        # row=ws.get_highest_row()
        # col=ws.get_highest_column()阿
        row = ws.max_row
        col = ws.max_column
        for i in range(2, row):
            # 获取楼幢的数据，ws是表，.rows是行，[i]是行数，[1]是这一行的第二列，因为list是从0开始的，.value用来取值。
            _buliding = ws.rows[i][1].value
            # 获取相对应字段的信息
            if _buliding == '龙北04':
                _buliding = '龙川北苑04南 '
            _room = str(ws.rows[i][2].value)
            _name = ws.rows[i][4].value
            _time = ws.rows[i][5].value
            if (_buliding and _room and _name and _time) is None:
                continue
            elif _room in list_room:
                a = list_room.index(_room)
                updata_name = list_a[a].get('name')
                updata_time = list_a[a].get('time')
                ''' 这边通过判断名字来确认序列，此代码因需求只需要判断两次，所以如果会出现一个人重复出现三次以上的情况的话会无法判断
                需要添加一个计数器来更改，逻辑上可行，但是我没测试过。'''
                if _name in list_name:
                    updata_name_list = updata_name.split(',')
                    if len(updata_name_list) <= 1:
                        updata_name = updata_name + '(两次)'
                        updata_time = updata_time + ',%s' % _time
                    else:
                        for onename in updata_name_list:
                            if _name == onename:
                                sy_updata_name = onename + '(两次)'
                                updata_time = updata_time + ',%s' % _time
                                # 这里是有bug的，如果按照一定的顺序发生的话时间会错乱。
                                sy_name = updata_name_list.index(onename)
                                # print(sy_name)
                            else:
                                pass
                        updata_name_list[sy_name] = sy_updata_name
                        updata_name = ','.join(updata_name_list)
                        # print(updata_name, updata_time)
                else:
                    updata_name = updata_name + ',%s' % _name
                    updata_time = updata_time + ',%s' % _time
                    # print(updata_name, updata_time)
                list_a[a]['name'] = updata_name
                list_a[a]['time'] = updata_time
                list_name.append(updata_name)
                continue
            else:
                dict_a = {'buliding': _buliding, 'room': _room,
                          'name': _name, 'time': _time}
            list_name.append(_name)
            list_room.append(_room)
            list_a.append(dict_a)
            # print(list_name, list_room, list_a, '\n\n')
        # print(len(list_room))
        # print(len(list_a))
        return list_a

    def load_b(self):
        one_row_data = {}
        self.lb04_data, self.tl07_data, self.sx08_data, self.xs04_data, self.lt03_data, self.tl08_data, self.xs01_data = [], [], [], [], [], [], []
        four = ['龙北04', '桃李07', '学士08', '学士04']
        three = ['鹿田03', '桃李08', '学士01']
        wb = load_workbook(filename=r'../b.xlsx')
        ws_all = wb.get_sheet_names()
        for i in ws_all:
            sheet_data = []
            # print(i)
            ws = wb.get_sheet_by_name(i)
            row = ws.max_row
            col = ws.max_column
            if i in four:
                # 执行取数据流程
                for one_row in range(2, row):
                    data = ws.rows[one_row]
                    level = data[8].value
                    problem = data[9].value
                    time = data[11].value
                    check = data[12].value
                    if problem is None:
                        problem = ''
                    if level is None:
                        level = ''
                    if time is None:
                        time = ''
                    if check is None:
                        check = ''
                    one_row_data = {'level': level, 'problem': problem,
                                    'time': time, 'check': check}
                    sheet_data.append(one_row_data)
                # 输出龙北04表内数据
                if i == '龙北04':
                    self.lb04_data = sheet_data
                # 输出桃李07表内数据
                elif i == '桃李07':
                    self.tl07_data = sheet_data
                # 输出学士08表内数据
                elif i == '学士08':
                    self.xs08_data = sheet_data
                # 输出学士04表内数据
                elif i == '学士04':
                    self.xs04_data = sheet_data
                else:
                    pass
            elif i in three:
                # 执行取数据流程
                for one_row in range(2, row):
                    data = ws.rows[one_row]
                    level = data[7].value
                    problem = data[8].value
                    time = data[10].value
                    check = data[11].value
                    if problem is None:
                        problem = ''
                    if level is None:
                        level = ''
                    if time is None:
                        time = ''
                    if check is None:
                        check = ''
                    one_row_data = {'level': level, 'problem': problem,
                                    'time': time, 'check': check}
                    sheet_data.append(one_row_data)
                # 输出鹿田03表内数据
                if i == '鹿田03':
                    self.lt03_data = sheet_data
                # 输出桃李08表内数据
                elif i == '桃李08':
                    self.tl08_data = sheet_data
                # 输出学士01表内数据
                elif i == '学士01':
                    self.xs01_data = sheet_data
                else:
                    pass
            else:
                continue

    def load_c(self):
        wb = load_workbook(filename=r'../c.xlsx')
        return wb

    def updata_a2c(self):
        ws_all = self.c.get_sheet_names()
        for i in self.list_a:
            # {'name': '周佳毅', 'buliding': '龙川北苑04南 ', 'class_name': '应电142', 'room': '217', 'time': '2016/6/7/22:01'}
            a_buliding_name = i.get('buliding')
            a_room = i.get('room')
            ws = self.c.get_sheet_by_name(a_buliding_name)
            row = ws.max_row
            col = ws.max_column
            # print(row,col) #行数和列数
            for lb in range(0, row):
                try:
                    if a_buliding_name == '龙川北苑04南 ':
                        c_room = ws.rows[lb][5].value[-3:]
                    elif a_buliding_name == '学士01':
                        c_room = ws.rows[lb][5].value
                    else:
                        continue
                    # print(c_room,a_room)
                    # print(type(c_room),type(a_room))  #<class 'str'> <class
                    # 'str'>
                    if int(c_room) == int(a_room):
                        # print('1')
                        name = i.get('name')
                        time = i.get('time')
                        # print(name, time)
                        ws.cell(row=lb + 1, column=8).value = '%s' % name
                        ws.cell(row=lb + 1, column=9).value = '%s' % time
                        # print('done')
                except:
                    continue

    def updata_b2c(self):
        ws_all = self.c.get_sheet_names()
        for i in ws_all[1:]:
            if i == '学士01':
                b_data = self.xs01_data
            elif i == '桃李08':
                b_data = self.tl08_data
            elif i == '鹿田南苑03幢':
                b_data = self.lt03_data
            elif i == '龙川北苑04南 ':
                b_data = self.lb04_data
            elif i == '学士苑04幢':
                b_data = self.xs04_data
            elif i == '学士苑08幢':
                b_data = self.xs08_data
            elif i == '桃李07':
                b_data = self.tl07_data
            else:
                continue
            ws = self.c.get_sheet_by_name(i)
            row = ws.max_row
            for (c_one_row, b_one_row) in zip(range(2, row), b_data):
                try:
                    ws.cell(row=c_one_row + 1,
                            column=10).value = '%s' % b_one_row.get('level')
                    ws.cell(row=c_one_row + 1,
                            column=11).value = '%s' % b_one_row.get('problem')
                    ws.cell(row=c_one_row + 1,
                            column=13).value = '%s' % b_one_row.get('time')
                    ws.cell(row=c_one_row + 1,
                            column=14).value = '%s' % b_one_row.get('check')
                except:
                    continue

    def updata_c2c(self):
        all_list = []
        sb = ['龙川北苑04南楼一楼','龙川北苑04南楼二楼','龙川北苑04南楼三楼','龙川北苑04南楼四楼','龙川北苑04南楼五楼']
        ws_all = self.c.get_sheet_names()
        # 将前面几张表格里的数据添加到汇总表格里
        for i in ws_all[1:8]:
            ws = self.c.get_sheet_by_name(i)
            row = ws.max_row
            for one_row in range(2, row):
                onerow_list = []
                data = ws.rows[one_row]
                if data[13].value is None:
                    continue
                else:
                    onerow_list.append(data[7].value)
                    onerow_list.append(data[8].value)
                    onerow_list.append(data[9].value)
                    onerow_list.append(data[10].value)
                    onerow_list.append(data[12].value)
                    onerow_list.append(data[13].value)
                    onerow_list.append(data[4].value)
                    onerow_list.append(data[5].value)
                    all_list.append(onerow_list)
        ws_huizong = self.c.get_sheet_by_name('汇总')
        huizong_row = ws_huizong.max_row
        for (huizong_one_row,onerow_data) in zip(range(2,huizong_row),all_list):
            for x in onerow_data:
                if x is None:
                    onerow_data[onerow_data.index(x)] = ''
            ws_huizong.cell(row=huizong_one_row + 1,
                    column=8).value = '%s' % onerow_data[0]
            ws_huizong.cell(row=huizong_one_row + 1,
                    column=9).value = '%s' % onerow_data[1]
            ws_huizong.cell(row=huizong_one_row + 1,
                    column=10).value = '%s' % onerow_data[2]
            ws_huizong.cell(row=huizong_one_row + 1,
                    column=11).value = '%s' % onerow_data[3]
            ws_huizong.cell(row=huizong_one_row + 1,
                    column=13).value = '%s' % onerow_data[4]
            ws_huizong.cell(row=huizong_one_row + 1,
                    column=14).value = '%s' % onerow_data[5]
        # 处理学生干部寝室部分
        ws_xsgb = self.c.get_sheet_by_name('学生干部寝室')
        ws_xsgb_row = ws_xsgb.max_row
        for i_wsgb in range(2,ws_xsgb_row):
            xsgb_data = ws_xsgb.rows[i_wsgb]
            xsgb_buliding = xsgb_data[4].value
            xsgb_room = str(xsgb_data[5].value)
            for seach in all_list:
                huizong_buliding = seach[6]
                huizong_room = str(seach[7])
                if huizong_buliding == '鹿田03':
                    huizong_buliding = '鹿田南苑03幢'
                elif huizong_buliding in sb:
                    huizong_buliding = '龙川北苑04南 '
                    # print(type(huizong_room))
                    huizong_room = huizong_room.split('-')[1]
                    # print(huizong_buliding,huizong_room)
                elif huizong_buliding == '学士苑08':
                    huizong_buliding = '学士苑08幢'
                if xsgb_buliding == huizong_buliding and xsgb_room == huizong_room:
                    for i_seach in seach:
                        if i_seach is None:
                            seach[seach.index(i_seach)] = ''
                        ws_xsgb.cell(row=i_wsgb + 1,
                                column=8).value = '%s' % seach[0]
                        ws_xsgb.cell(row=i_wsgb + 1,
                                column=9).value = '%s' % seach[1]
                        ws_xsgb.cell(row=i_wsgb + 1,
                                column=10).value = '%s' % seach[2]
                        ws_xsgb.cell(row=i_wsgb + 1,
                                column=11).value = '%s' % seach[3]

    def main(self):
        self.updata_a2c()
        self.updata_b2c()
        self.updata_c2c()
        self.c.save('../c.xlsx')


if __name__ == "__main__":
    ExcelSGH()
